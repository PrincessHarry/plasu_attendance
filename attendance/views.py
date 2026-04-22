import csv
import json
import secrets
import hashlib
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.conf import settings
from django.db.models import Count, Q

from .utils import get_client_ip, get_device_info, generate_session_token, build_qr_payload, session_expires_at
from .models import (
    User, Faculty, Department, Course, LecturerProfile,
    StudentProfile, FingerprintTemplate, AttendanceSession, AttendanceRecord
)
from .decorators import role_required


# ─── AUTH ────────────────────────────────────────────────────────────────────

def index_redirect(request):
    if request.user.is_authenticated:
        return redirect('attendance:dashboard')
    return redirect('attendance:login')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('attendance:dashboard')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, email=email, password=password)
        if user:
            if user.is_active:
                login(request, user)
                next_url = request.GET.get('next', '')
                if next_url:
                    return redirect(next_url)
                return redirect('attendance:dashboard')
            else:
                messages.error(request, 'Your account has been deactivated.')
        else:
            messages.error(request, 'Invalid email or password.')

    return render(request, 'attendance/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out.')
    return redirect('attendance:login')


@login_required
def dashboard_redirect(request):
    if request.user.role == 'admin':
        return redirect('attendance:admin_dashboard')
    elif request.user.role == 'lecturer':
        return redirect('attendance:lecturer_dashboard')
    else:
        return redirect('attendance:student_dashboard')


# ─── ADMIN VIEWS ─────────────────────────────────────────────────────────────

@login_required
@role_required('admin')
def admin_dashboard(request):
    context = {
        'total_students': StudentProfile.objects.count(),
        'total_lecturers': LecturerProfile.objects.count(),
        'total_courses': Course.objects.count(),
        'total_faculties': Faculty.objects.count(),
        'total_departments': Department.objects.count(),
        'active_sessions': AttendanceSession.objects.filter(status='active').count(),
        'recent_sessions': AttendanceSession.objects.select_related('course', 'lecturer__user').order_by('-started_at')[:5],
        'recent_students': StudentProfile.objects.select_related('user', 'department').order_by('-created_at')[:5],
    }
    return render(request, 'attendance/admin/dashboard.html', context)


@login_required
@role_required('admin')
def manage_faculties(request):
    faculties = Faculty.objects.annotate(dept_count=Count('departments')).order_by('name')
    return render(request, 'attendance/admin/faculties.html', {'faculties': faculties})


@login_required
@role_required('admin')
def add_faculty(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        description = request.POST.get('description', '').strip()
        if name and code:
            if Faculty.objects.filter(Q(name=name) | Q(code=code)).exists():
                messages.error(request, 'Faculty with this name or code already exists.')
            else:
                Faculty.objects.create(name=name, code=code, description=description)
                messages.success(request, f'Faculty "{name}" created successfully.')
                return redirect('attendance:manage_faculties')
        else:
            messages.error(request, 'Name and code are required.')
    return render(request, 'attendance/admin/faculty_form.html', {'action': 'Add'})


@login_required
@role_required('admin')
def edit_faculty(request, pk):
    faculty = get_object_or_404(Faculty, pk=pk)
    if request.method == 'POST':
        faculty.name = request.POST.get('name', faculty.name).strip()
        faculty.code = request.POST.get('code', faculty.code).strip().upper()
        faculty.description = request.POST.get('description', '').strip()
        faculty.save()
        messages.success(request, 'Faculty updated successfully.')
        return redirect('attendance:manage_faculties')
    return render(request, 'attendance/admin/faculty_form.html', {'action': 'Edit', 'faculty': faculty})


@login_required
@role_required('admin')
def delete_faculty(request, pk):
    faculty = get_object_or_404(Faculty, pk=pk)
    if request.method == 'POST':
        name = faculty.name
        faculty.delete()
        messages.success(request, f'Faculty "{name}" deleted.')
    return redirect('attendance:manage_faculties')


@login_required
@role_required('admin')
def manage_departments(request):
    departments = Department.objects.select_related('faculty').annotate(
        student_count=Count('students'), course_count=Count('courses')
    ).order_by('faculty__name', 'name')
    faculties = Faculty.objects.all()
    return render(request, 'attendance/admin/departments.html', {
        'departments': departments, 'faculties': faculties
    })


@login_required
@role_required('admin')
def add_department(request):
    faculties = Faculty.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        faculty_id = request.POST.get('faculty')
        description = request.POST.get('description', '').strip()
        if name and code and faculty_id:
            faculty = get_object_or_404(Faculty, pk=faculty_id)
            if Department.objects.filter(code=code).exists():
                messages.error(request, 'Department code already exists.')
            else:
                Department.objects.create(name=name, code=code, faculty=faculty, description=description)
                messages.success(request, f'Department "{name}" created.')
                return redirect('attendance:manage_departments')
        else:
            messages.error(request, 'All fields are required.')
    return render(request, 'attendance/admin/department_form.html', {'action': 'Add', 'faculties': faculties})


@login_required
@role_required('admin')
def edit_department(request, pk):
    department = get_object_or_404(Department, pk=pk)
    faculties = Faculty.objects.all()
    if request.method == 'POST':
        department.name = request.POST.get('name', department.name).strip()
        department.code = request.POST.get('code', department.code).strip().upper()
        faculty_id = request.POST.get('faculty')
        if faculty_id:
            department.faculty = get_object_or_404(Faculty, pk=faculty_id)
        department.description = request.POST.get('description', '').strip()
        department.save()
        messages.success(request, 'Department updated.')
        return redirect('attendance:manage_departments')
    return render(request, 'attendance/admin/department_form.html', {
        'action': 'Edit', 'department': department, 'faculties': faculties
    })


@login_required
@role_required('admin')
def delete_department(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        department.delete()
        messages.success(request, 'Department deleted.')
    return redirect('attendance:manage_departments')


@login_required
@role_required('admin')
def manage_courses(request):
    courses = Course.objects.select_related('department', 'lecturer__user').annotate(
        student_count=Count('students')
    ).order_by('code')
    return render(request, 'attendance/admin/courses.html', {'courses': courses})


@login_required
@role_required('admin')
def add_course(request):
    departments = Department.objects.select_related('faculty').all()
    lecturers = LecturerProfile.objects.select_related('user').all()
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        title = request.POST.get('title', '').strip()
        dept_id = request.POST.get('department')
        lecturer_id = request.POST.get('lecturer')
        credit_units = request.POST.get('credit_units', 3)
        semester = request.POST.get('semester', 'first')
        level = request.POST.get('level', 100)
        if code and title and dept_id:
            if Course.objects.filter(code=code).exists():
                messages.error(request, 'Course code already exists.')
            else:
                dept = get_object_or_404(Department, pk=dept_id)
                lecturer = LecturerProfile.objects.filter(pk=lecturer_id).first() if lecturer_id else None
                Course.objects.create(
                    code=code, title=title, department=dept, lecturer=lecturer,
                    credit_units=credit_units, semester=semester, level=level
                )
                messages.success(request, f'Course "{code}" created.')
                return redirect('attendance:manage_courses')
        else:
            messages.error(request, 'Code, title, and department are required.')
    return render(request, 'attendance/admin/course_form.html', {
        'action': 'Add', 'departments': departments, 'lecturers': lecturers
    })


@login_required
@role_required('admin')
def edit_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    departments = Department.objects.select_related('faculty').all()
    lecturers = LecturerProfile.objects.select_related('user').all()
    if request.method == 'POST':
        course.code = request.POST.get('code', course.code).strip().upper()
        course.title = request.POST.get('title', course.title).strip()
        dept_id = request.POST.get('department')
        lecturer_id = request.POST.get('lecturer')
        if dept_id:
            course.department = get_object_or_404(Department, pk=dept_id)
        course.lecturer = LecturerProfile.objects.filter(pk=lecturer_id).first() if lecturer_id else None
        course.credit_units = request.POST.get('credit_units', course.credit_units)
        course.semester = request.POST.get('semester', course.semester)
        course.level = request.POST.get('level', course.level)
        course.save()
        messages.success(request, 'Course updated.')
        return redirect('attendance:manage_courses')
    return render(request, 'attendance/admin/course_form.html', {
        'action': 'Edit', 'course': course, 'departments': departments, 'lecturers': lecturers
    })


@login_required
@role_required('admin')
def delete_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.delete()
        messages.success(request, 'Course deleted.')
    return redirect('attendance:manage_courses')


@login_required
@role_required('admin')
def manage_lecturers(request):
    lecturers = LecturerProfile.objects.select_related('user', 'department__faculty').annotate(
        course_count=Count('courses')
    ).order_by('user__last_name')
    return render(request, 'attendance/admin/lecturers.html', {'lecturers': lecturers})


@login_required
@role_required('admin')
def add_lecturer(request):
    departments = Department.objects.select_related('faculty').all()
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        staff_id = request.POST.get('staff_id', '').strip()
        dept_id = request.POST.get('department')
        phone = request.POST.get('phone', '').strip()
        qualification = request.POST.get('qualification', '').strip()
        if all([first_name, last_name, email, password, staff_id]):
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already exists.')
            elif LecturerProfile.objects.filter(staff_id=staff_id).exists():
                messages.error(request, 'Staff ID already exists.')
            else:
                user = User.objects.create_user(
                    email=email, password=password,
                    first_name=first_name, last_name=last_name, role='lecturer'
                )
                dept = Department.objects.filter(pk=dept_id).first() if dept_id else None
                LecturerProfile.objects.create(
                    user=user, staff_id=staff_id, department=dept,
                    phone=phone, qualification=qualification
                )
                # Fingerprint is auto-created by signal; update reference if needed
                FingerprintTemplate.objects.get_or_create(
                    user=user,
                    defaults={
                        'template_reference': f"LECT-FP-{staff_id}",
                        'template_hash': '',  # Will be set during first enrollment
                    }
                )
                messages.success(request, f'Lecturer "{first_name} {last_name}" created.')
                return redirect('attendance:manage_lecturers')
        else:
            messages.error(request, 'All required fields must be filled.')
    return render(request, 'attendance/admin/lecturer_form.html', {'action': 'Add', 'departments': departments})


@login_required
@role_required('admin')
def edit_lecturer(request, pk):
    lecturer = get_object_or_404(LecturerProfile, pk=pk)
    departments = Department.objects.select_related('faculty').all()
    if request.method == 'POST':
        lecturer.user.first_name = request.POST.get('first_name', lecturer.user.first_name).strip()
        lecturer.user.last_name = request.POST.get('last_name', lecturer.user.last_name).strip()
        lecturer.user.save()
        lecturer.staff_id = request.POST.get('staff_id', lecturer.staff_id).strip()
        dept_id = request.POST.get('department')
        lecturer.department = Department.objects.filter(pk=dept_id).first() if dept_id else None
        lecturer.phone = request.POST.get('phone', '').strip()
        lecturer.qualification = request.POST.get('qualification', '').strip()
        lecturer.save()
        messages.success(request, 'Lecturer updated.')
        return redirect('attendance:manage_lecturers')
    return render(request, 'attendance/admin/lecturer_form.html', {
        'action': 'Edit', 'lecturer': lecturer, 'departments': departments
    })


@login_required
@role_required('admin')
def delete_lecturer(request, pk):
    lecturer = get_object_or_404(LecturerProfile, pk=pk)
    if request.method == 'POST':
        lecturer.user.delete()
        messages.success(request, 'Lecturer deleted.')
    return redirect('attendance:manage_lecturers')


@login_required
@role_required('admin')
def manage_students(request):
    students = StudentProfile.objects.select_related('user', 'department__faculty').order_by('user__last_name')
    return render(request, 'attendance/admin/students.html', {'students': students})


@login_required
@role_required('admin')
def add_student(request):
    departments = Department.objects.select_related('faculty').all()
    courses = Course.objects.select_related('department').all()
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        matric = request.POST.get('matric_number', '').strip()
        dept_id = request.POST.get('department')
        level = request.POST.get('level', 100)
        phone = request.POST.get('phone', '').strip()
        if all([first_name, last_name, email, password, matric]):
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already exists.')
            elif StudentProfile.objects.filter(matric_number=matric).exists():
                messages.error(request, 'Matric number already exists.')
            else:
                user = User.objects.create_user(
                    email=email, password=password,
                    first_name=first_name, last_name=last_name, role='student'
                )
                dept = Department.objects.filter(pk=dept_id).first() if dept_id else None
                student = StudentProfile.objects.create(
                    user=user, matric_number=matric, department=dept,
                    level=level, phone=phone
                )
                # Enroll in courses
                course_ids = request.POST.getlist('courses')
                if course_ids:
                    student.courses.set(Course.objects.filter(pk__in=course_ids))
                # Fingerprint is auto-created by signal; update reference if needed
                FingerprintTemplate.objects.get_or_create(
                    user=user,
                    defaults={
                        'template_reference': f"STU-FP-{matric.replace('/', '-')}",
                        'template_hash': '',  # Will be set during first enrollment
                    }
                )
                messages.success(request, f'Student "{first_name} {last_name}" registered.')
                return redirect('attendance:manage_students')
        else:
            messages.error(request, 'All required fields must be filled.')
    return render(request, 'attendance/admin/student_form.html', {
        'action': 'Add', 'departments': departments, 'courses': courses
    })


@login_required
@role_required('admin')
def edit_student(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    departments = Department.objects.select_related('faculty').all()
    courses = Course.objects.select_related('department').all()
    if request.method == 'POST':
        student.user.first_name = request.POST.get('first_name', student.user.first_name).strip()
        student.user.last_name = request.POST.get('last_name', student.user.last_name).strip()
        student.user.save()
        student.matric_number = request.POST.get('matric_number', student.matric_number).strip()
        dept_id = request.POST.get('department')
        student.department = Department.objects.filter(pk=dept_id).first() if dept_id else None
        student.level = request.POST.get('level', student.level)
        student.phone = request.POST.get('phone', '').strip()
        student.save()
        course_ids = request.POST.getlist('courses')
        student.courses.set(Course.objects.filter(pk__in=course_ids))
        messages.success(request, 'Student updated.')
        return redirect('attendance:manage_students')
    return render(request, 'attendance/admin/student_form.html', {
        'action': 'Edit', 'student': student, 'departments': departments, 'courses': courses
    })


@login_required
@role_required('admin')
def delete_student(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    if request.method == 'POST':
        student.user.delete()
        messages.success(request, 'Student deleted.')
    return redirect('attendance:manage_students')


@login_required
@role_required('admin')
def admin_attendance_records(request):
    records = AttendanceRecord.objects.select_related(
        'session__course', 'student__user', 'session__lecturer__user'
    ).order_by('-marked_at')
    courses = Course.objects.all()
    # Filter
    course_id = request.GET.get('course')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if course_id:
        records = records.filter(session__course_id=course_id)
    if date_from:
        records = records.filter(marked_at__date__gte=date_from)
    if date_to:
        records = records.filter(marked_at__date__lte=date_to)
    return render(request, 'attendance/admin/attendance_records.html', {
        'records': records[:200], 'courses': courses
    })


@login_required
@role_required('admin')
def export_attendance_csv(request):
    records = AttendanceRecord.objects.select_related(
        'session__course', 'student__user', 'session__lecturer__user'
    ).order_by('-marked_at')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
    writer = csv.writer(response)
    from .utils import attendance_csv_rows
    for row in attendance_csv_rows(records):
        writer.writerow(row)
    return response


# ─── LECTURER VIEWS ──────────────────────────────────────────────────────────

@login_required
@role_required('lecturer')
def lecturer_dashboard(request):
    lecturer = get_object_or_404(LecturerProfile, user=request.user)
    courses = Course.objects.filter(lecturer=lecturer).annotate(student_count=Count('students'))
    active_sessions = AttendanceSession.objects.filter(
        lecturer=lecturer, status='active'
    ).select_related('course')
    recent_sessions = AttendanceSession.objects.filter(
        lecturer=lecturer
    ).select_related('course').order_by('-started_at')[:5]
    context = {
        'lecturer': lecturer,
        'courses': courses,
        'active_sessions': active_sessions,
        'recent_sessions': recent_sessions,
        'total_sessions': AttendanceSession.objects.filter(lecturer=lecturer).count(),
        'total_students': StudentProfile.objects.filter(courses__lecturer=lecturer).distinct().count(),
    }
    return render(request, 'attendance/lecturer/dashboard.html', context)


@login_required
@role_required('lecturer')
def lecturer_courses(request):
    lecturer = get_object_or_404(LecturerProfile, user=request.user)
    courses = Course.objects.filter(lecturer=lecturer).select_related('department').annotate(
        student_count=Count('students')
    )
    return render(request, 'attendance/lecturer/courses.html', {'courses': courses, 'lecturer': lecturer})


@login_required
@role_required('lecturer')
def lecturer_sessions(request):
    lecturer = get_object_or_404(LecturerProfile, user=request.user)
    sessions = AttendanceSession.objects.filter(lecturer=lecturer).select_related('course').annotate(
        present_count=Count('records', filter=Q(records__status='present'))
    ).order_by('-started_at')
    return render(request, 'attendance/lecturer/sessions.html', {'sessions': sessions})


@login_required
@role_required('lecturer')
def start_session(request):
    lecturer = get_object_or_404(LecturerProfile, user=request.user)
    courses = Course.objects.filter(lecturer=lecturer)
    if request.method == 'POST':
        course_id = request.POST.get('course')
        venue = request.POST.get('venue', '').strip()
        duration = int(request.POST.get('duration', 60))  # minutes
        if course_id:
            course = get_object_or_404(Course, pk=course_id, lecturer=lecturer)
            # Generate session token
            token = secrets.token_urlsafe(32)
            expires_at = timezone.now() + timedelta(minutes=duration)
            qr_data = json.dumps({
                'session_token': token,
                'course_code': course.code,
                'expires_at': expires_at.isoformat(),
            })
            session = AttendanceSession.objects.create(
                course=course,
                lecturer=lecturer,
                session_token=token,
                qr_data=qr_data,
                expires_at=expires_at,
                venue=venue
            )
            messages.success(request, f'Session started for {course.code}')
            return redirect('attendance:session_detail', session_id=session.id)
    return render(request, 'attendance/lecturer/start_session.html', {'courses': courses})


@login_required
@role_required('lecturer')
def session_detail(request, session_id):
    lecturer = get_object_or_404(LecturerProfile, user=request.user)
    session = get_object_or_404(AttendanceSession, pk=session_id, lecturer=lecturer)
    records = AttendanceRecord.objects.filter(session=session).select_related('student__user').order_by('-marked_at')
    attend_url = request.build_absolute_uri(f'/attend/{session.session_token}/')
    context = {
        'session': session,
        'records': records,
        'attend_url': attend_url,
        'enrolled_count': session.course.students.count(),
    }
    return render(request, 'attendance/lecturer/session_detail.html', context)


@login_required
@role_required('lecturer')
def session_qr(request, session_id):
    lecturer = get_object_or_404(LecturerProfile, user=request.user)
    session = get_object_or_404(AttendanceSession, pk=session_id, lecturer=lecturer)
    attend_url = request.build_absolute_uri(f'/attend/{session.session_token}/')
    return render(request, 'attendance/lecturer/session_qr.html', {
        'session': session, 'attend_url': attend_url
    })


@login_required
@role_required('lecturer')
def end_session(request, session_id):
    lecturer = get_object_or_404(LecturerProfile, user=request.user)
    session = get_object_or_404(AttendanceSession, pk=session_id, lecturer=lecturer)
    if request.method == 'POST':
        session.status = 'ended'
        session.ended_at = timezone.now()
        session.save()
        messages.success(request, 'Session ended successfully.')
    return redirect('attendance:session_detail', session_id=session.id)


@login_required
@role_required('lecturer')
def lecturer_attendance_history(request, course_id):
    lecturer = get_object_or_404(LecturerProfile, user=request.user)
    course = get_object_or_404(Course, pk=course_id, lecturer=lecturer)
    sessions = AttendanceSession.objects.filter(
        course=course, lecturer=lecturer
    ).annotate(present_count=Count('records', filter=Q(records__status='present'))).order_by('-started_at')
    return render(request, 'attendance/lecturer/attendance_history.html', {
        'course': course, 'sessions': sessions
    })


# ─── STUDENT VIEWS ───────────────────────────────────────────────────────────

@login_required
@role_required('student')
def student_dashboard(request):
    student = get_object_or_404(StudentProfile, user=request.user)
    courses = student.courses.select_related('department', 'lecturer__user').all()
    records = AttendanceRecord.objects.filter(student=student).select_related(
        'session__course'
    ).order_by('-marked_at')[:10]

    # Stats per course
    course_stats = []
    for course in courses:
        total = AttendanceSession.objects.filter(course=course, status='ended').count()
        attended = AttendanceRecord.objects.filter(student=student, session__course=course, status='present').count()
        pct = round((attended / total * 100) if total > 0 else 0)
        course_stats.append({
            'course': course,
            'total_sessions': total,
            'attended': attended,
            'percentage': pct,
        })

    context = {
        'student': student,
        'courses': courses,
        'records': records,
        'course_stats': course_stats,
        'total_present': AttendanceRecord.objects.filter(student=student, status='present').count(),
    }
    return render(request, 'attendance/student/dashboard.html', context)


@login_required
@role_required('student')
def student_attendance_history(request):
    student = get_object_or_404(StudentProfile, user=request.user)
    records = AttendanceRecord.objects.filter(student=student).select_related(
        'session__course', 'session__lecturer__user'
    ).order_by('-marked_at')
    return render(request, 'attendance/student/attendance_history.html', {
        'student': student, 'records': records
    })


@login_required
@role_required('student')
def scan_qr(request):
    return render(request, 'attendance/student/scan_qr.html')


def attend_session(request, token):
    """Session attendance page - accessible via QR code link"""
    session = get_object_or_404(AttendanceSession, session_token=token)

    # Check session validity
    if not session.is_active:
        return render(request, 'attendance/attend_error.html', {
            'error': 'This attendance session has expired or ended.',
            'session': session
        })

    # Must be logged in as a student
    if not request.user.is_authenticated:
        return redirect(f'/login/?next=/attend/{token}/')

    if request.user.role != 'student':
        return render(request, 'attendance/attend_error.html', {
            'error': 'Only students can mark attendance.',
        })

    student = get_object_or_404(StudentProfile, user=request.user)

    # Check if already marked
    already_marked = AttendanceRecord.objects.filter(session=session, student=student).exists()
    if already_marked:
        record = AttendanceRecord.objects.get(session=session, student=student)
        return render(request, 'attendance/attend_success.html', {
            'session': session, 'student': student, 'record': record, 'already_done': True
        })

    # Check if student is enrolled
    enrolled = session.course.students.filter(pk=student.pk).exists()

    context = {
        'session': session,
        'student': student,
        'enrolled': enrolled,
        'has_fingerprint': FingerprintTemplate.objects.filter(user=request.user, is_active=True).exists(),
    }
    return render(request, 'attendance/attend_session.html', context)


@login_required
def verify_fingerprint(request, session_id):
    session = get_object_or_404(AttendanceSession, pk=session_id)
    student = get_object_or_404(StudentProfile, user=request.user)
    if request.method == 'POST':
        # Get fingerprint data from request
        fp_input = request.POST.get('fingerprint_data', '')
        fingerprint = FingerprintTemplate.objects.filter(user=request.user, is_active=True).first()
        if fingerprint:
            # Check for duplicate fingerprints across all users
            duplicate_fingerprint = FingerprintTemplate.objects.filter(
                template_hash=fp_input,
                is_active=True
            ).exclude(user=request.user).first()

            if duplicate_fingerprint:
                # Flag the duplicate immediately
                duplicate_student = StudentProfile.objects.filter(user=duplicate_fingerprint.user).first()
                duplicate_name = duplicate_student.user.get_full_name() if duplicate_student else "another student"
                duplicate_matric = duplicate_student.matric_number if duplicate_student else "unknown"

                # Log the security incident
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"FINGERPRINT DUPLICATE DETECTED: User {request.user.get_full_name()} ({student.matric_number}) attempted to use fingerprint already enrolled for {duplicate_name} ({duplicate_matric})")

                return JsonResponse({
                    'status': 'error',
                    'message': f'Security Alert: This fingerprint is already registered to {duplicate_name}. Contact administrator immediately.',
                    'duplicate_detected': True
                })

            # Simulated verification: hash the input and compare
            verified = True  # In production: compare actual template
            if verified:
                # Update fingerprint template with new hash
                fingerprint.template_hash = fp_input
                fingerprint.save()

                if not AttendanceRecord.objects.filter(session=session, student=student).exists():
                    AttendanceRecord.objects.create(
                        session=session, student=student, status='present',
                        fingerprint_verified=True,
                        ip_address=request.META.get('REMOTE_ADDR'),
                        device_info=request.META.get('HTTP_USER_AGENT', '')[:500]
                    )
                return JsonResponse({'status': 'success', 'message': 'Attendance marked!'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Fingerprint verification failed.'})
        return JsonResponse({'status': 'error', 'message': 'No fingerprint enrolled.'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


# ─── API ENDPOINTS ────────────────────────────────────────────────────────────

@login_required
def api_session_status(request, session_id):
    session = get_object_or_404(AttendanceSession, pk=session_id)
    return JsonResponse({
        'status': session.status,
        'is_active': session.is_active,
        'attendance_count': session.attendance_count,
        'expires_at': session.expires_at.isoformat(),
    })


@login_required
def api_attendance_count(request, session_id):
    session = get_object_or_404(AttendanceSession, pk=session_id)
    records = AttendanceRecord.objects.filter(session=session, status='present').select_related(
        'student__user'
    ).order_by('-marked_at')[:20]
    return JsonResponse({
        'count': records.count(),
        'enrolled': session.course.students.count(),
        'students': [
            {
                'name': r.student.user.get_full_name(),
                'matric': r.student.matric_number,
                'time': r.marked_at.strftime('%H:%M:%S'),
                'verified': r.fingerprint_verified,
            } for r in records
        ]
    })


@login_required
@require_POST
def api_verify_fingerprint(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        fingerprint_data = data.get('fingerprint_data')
        session = get_object_or_404(AttendanceSession, pk=session_id)

        if not session.is_active:
            return JsonResponse({'status': 'error', 'message': 'Session is not active.'})

        if request.user.role != 'student':
            return JsonResponse({'status': 'error', 'message': 'Only students can mark attendance.'})

        student = get_object_or_404(StudentProfile, user=request.user)

        if AttendanceRecord.objects.filter(session=session, student=student).exists():
            return JsonResponse({'status': 'error', 'message': 'Attendance already marked.'})

        fingerprint = FingerprintTemplate.objects.filter(user=request.user, is_active=True).first()
        if not fingerprint:
            return JsonResponse({'status': 'error', 'message': 'No fingerprint enrolled for this account.'})

        # Check if this fingerprint data is already enrolled for another student
        duplicate_fingerprint = FingerprintTemplate.objects.filter(
            template_hash=fingerprint_data,
            is_active=True
        ).exclude(user=request.user).first()

        if duplicate_fingerprint:
            # Flag the duplicate immediately - this is a security violation
            duplicate_student = StudentProfile.objects.filter(user=duplicate_fingerprint.user).first()
            duplicate_name = duplicate_student.user.get_full_name() if duplicate_student else "another student"
            duplicate_matric = duplicate_student.matric_number if duplicate_student else "unknown"

            # Log the security incident
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"FINGERPRINT DUPLICATE DETECTED: User {request.user.get_full_name()} ({student.matric_number}) attempted to use fingerprint already enrolled for {duplicate_name} ({duplicate_matric})")

            return JsonResponse({
                'status': 'error',
                'message': f'Security Alert: This fingerprint is already registered to {duplicate_name}. Contact administrator immediately.',
                'duplicate_detected': True
            })

        # Check if fingerprint matches the enrolled template (simulated comparison)
        # In production, this would compare actual biometric data
        fingerprint_match = True  # Simulated match

        if not fingerprint_match:
            return JsonResponse({'status': 'error', 'message': 'Fingerprint verification failed. Please try again.'})

        # Update the fingerprint template with the new hash for future verification
        fingerprint.template_hash = fingerprint_data
        fingerprint.save()

        # Mark attendance
        AttendanceRecord.objects.create(
            session=session, student=student, status='present',
            fingerprint_verified=True,
            ip_address=request.META.get('REMOTE_ADDR'),
            device_info=request.META.get('HTTP_USER_AGENT', '')[:500]
        )
        return JsonResponse({
            'status': 'success',
            'message': f'Attendance marked for {student.user.get_full_name()}',
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required
@require_POST
def api_mark_attendance(request):
    """Mark attendance after successful fingerprint verification"""
    try:
        data = json.loads(request.body)
        token = data.get('token')
        session = get_object_or_404(AttendanceSession, session_token=token)

        if not session.is_active:
            return JsonResponse({'success': False, 'message': 'Session expired or ended.'})

        student = get_object_or_404(StudentProfile, user=request.user)

        if AttendanceRecord.objects.filter(session=session, student=student).exists():
            return JsonResponse({'success': False, 'message': 'Attendance already recorded.'})

        AttendanceRecord.objects.create(
            session=session, student=student, status='present',
            fingerprint_verified=True,
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        return JsonResponse({'success': True, 'message': 'Attendance recorded successfully!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ─── ERROR HANDLERS ───────────────────────────────────────────────────────────

def error_404(request, exception=None):
    return render(request, '404.html', status=404)


def error_500(request):
    return render(request, '500.html', status=500)


# ─── BULK STUDENT UPLOAD ─────────────────────────────────────────────────────

@login_required
@role_required('admin')
def bulk_upload_students(request):
    """Upload an Excel/CSV spreadsheet to register multiple students at once."""
    departments = Department.objects.select_related('faculty').all()
    courses     = Course.objects.select_related('department').all()

    if request.method == 'POST':
        upload_file = request.FILES.get('spreadsheet')
        if not upload_file:
            messages.error(request, 'No file selected.')
            return redirect('attendance:bulk_upload_students')

        filename = upload_file.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                import io
                decoded = upload_file.read().decode('utf-8-sig')
                reader  = csv.DictReader(io.StringIO(decoded))
                rows    = list(reader)

            elif filename.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                    from openpyxl import load_workbook
                    wb   = load_workbook(filename=upload_file, read_only=True, data_only=True)
                    ws   = wb.active
                    hdrs = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(hdrs, row)))
                except ImportError:
                    messages.error(request, 'openpyxl is required for Excel files. Run: pip install openpyxl')
                    return redirect('attendance:bulk_upload_students')
            else:
                messages.error(request, 'Unsupported file type. Upload a .csv or .xlsx file.')
                return redirect('attendance:bulk_upload_students')

        except Exception as e:
            messages.error(request, f'Could not read file: {e}')
            return redirect('attendance:bulk_upload_students')

        created_count = 0
        skipped_count = 0
        errors        = []

        for i, row in enumerate(rows, start=2):
            # Normalise keys (strip whitespace, lowercase)
            row = {k.strip().lower().replace(' ', '_'): (str(v).strip() if v is not None else '') for k, v in row.items()}

            first_name   = row.get('first_name', '').strip()
            last_name    = row.get('last_name',  '').strip()
            email        = row.get('email',       '').strip().lower()
            matric       = row.get('matric_number', row.get('matric', '')).strip()
            dept_code    = row.get('department_code', row.get('department', '')).strip().upper()
            level_raw    = row.get('level', '100')
            phone        = row.get('phone', '')

            # Derive email if missing
            if not email and first_name and last_name:
                email = f"{first_name.split()[0].lower()}.{last_name.lower()}@student.plasu.edu.ng"

            if not all([first_name, last_name, email, matric]):
                errors.append(f"Row {i}: missing required fields (first_name, last_name, email, matric_number).")
                continue

            if User.objects.filter(email=email).exists():
                skipped_count += 1
                continue

            if StudentProfile.objects.filter(matric_number=matric).exists():
                skipped_count += 1
                continue

            # Resolve level
            try:
                level = int(str(level_raw).replace('L', '').replace('l', '').strip())
            except (ValueError, TypeError):
                level = 100

            # Resolve department
            dept = Department.objects.filter(code=dept_code).first()

            # Generate password
            serial   = matric.split('/')[-1] if '/' in matric else matric[-4:]
            password = f"PLASUStu@{serial}"

            try:
                user = User.objects.create_user(
                    email=email, password=password,
                    first_name=first_name, last_name=last_name, role='student'
                )
                student = StudentProfile.objects.create(
                    user=user, department=dept,
                    matric_number=matric, level=level, phone=phone
                )

                # Enroll in courses specified in the row (comma-separated codes)
                course_codes_str = row.get('courses', row.get('course_codes', ''))
                if course_codes_str:
                    for code in [c.strip().upper() for c in course_codes_str.split(',') if c.strip()]:
                        course_obj = Course.objects.filter(code=code).first()
                        if course_obj:
                            student.courses.add(course_obj)

                # Send welcome email
                _send_welcome_email(user, password, matric)

                created_count += 1

            except Exception as e:
                errors.append(f"Row {i} ({email}): {e}")
                continue

        msg_parts = [f'{created_count} student(s) registered successfully.']
        if skipped_count:
            msg_parts.append(f'{skipped_count} skipped (already exist).')
        if errors:
            msg_parts.append(f'{len(errors)} error(s).')

        if created_count:
            messages.success(request, ' '.join(msg_parts))
        else:
            messages.warning(request, ' '.join(msg_parts))

        if errors:
            for err in errors[:5]:   # show first 5 only
                messages.error(request, err)

        return redirect('attendance:manage_students')

    default_cols = [
        ('first_name',    'Student first name'),
        ('last_name',     'Student last name (surname)'),
        ('email',         'Login email — auto-generated if blank'),
        ('matric_number', 'e.g. PLASU/2023/FNAS/001'),
    ]
    return render(request, 'attendance/admin/bulk_upload_students.html', {
        'departments':   departments,
        'courses':       courses,
        'default_cols':  default_cols,
    })


def _send_welcome_email(user, password, matric):
    """Send login credentials to a newly registered student."""
    from django.core.mail import send_mail
    from django.conf import settings as django_settings
    subject = 'Your PLASU Smart Attendance Account'
    body = (
        f"Dear {user.get_full_name()},\n\n"
        f"Your PLASU Smart Attendance System account has been created.\n\n"
        f"Login Details:\n"
        f"  Email    : {user.email}\n"
        f"  Password : {password}\n"
        f"  Matric   : {matric}\n\n"
        f"Please log in at http://127.0.0.1:8000/login/ and change your password.\n\n"
        f"Plateau State University\n"
        f"Smart Attendance System\n"
    )
    try:
        send_mail(
            subject, body,
            django_settings.DEFAULT_FROM_EMAIL,
            [user.email],
            fail_silently=True,
        )
    except Exception:
        pass   # Email failure must not interrupt bulk upload


# ─── PASSWORD RESET (simple token-based) ─────────────────────────────────────

def forgot_password(request):
    """Step 1 — collect email and send reset link."""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        user  = User.objects.filter(email=email).first()
        # Always show the same message to prevent email enumeration
        if user:
            _send_password_reset_email(request, user)
        messages.success(
            request,
            'If that email is registered, a password reset link has been sent. Check your inbox.'
        )
        return redirect('attendance:forgot_password')
    return render(request, 'attendance/forgot_password.html')


def reset_password(request, token):
    """Step 2 — validate token and set new password."""
    from django.core.cache import cache
    user_id = cache.get(f'pwd_reset_{token}')
    if not user_id:
        messages.error(request, 'This reset link is invalid or has expired.')
        return redirect('attendance:forgot_password')

    user = User.objects.filter(pk=user_id).first()
    if not user:
        messages.error(request, 'User not found.')
        return redirect('attendance:forgot_password')

    if request.method == 'POST':
        new_pwd  = request.POST.get('password', '')
        confirm  = request.POST.get('confirm_password', '')
        if len(new_pwd) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
        elif new_pwd != confirm:
            messages.error(request, 'Passwords do not match.')
        else:
            user.set_password(new_pwd)
            user.save()
            cache.delete(f'pwd_reset_{token}')
            messages.success(request, 'Password changed successfully. You can now sign in.')
            return redirect('attendance:login')

    return render(request, 'attendance/reset_password.html', {'token': token})


def _send_password_reset_email(request, user):
    """Generate a reset token, cache it for 1 hour, and email the link."""
    import secrets as _secrets
    from django.core.cache import cache
    from django.core.mail import send_mail
    from django.conf import settings as django_settings

    token    = _secrets.token_urlsafe(32)
    cache.set(f'pwd_reset_{token}', str(user.pk), timeout=3600)
    reset_url = request.build_absolute_uri(f'/reset-password/{token}/')

    subject = 'PLASU Attendance — Password Reset'
    body    = (
        f"Dear {user.get_full_name()},\n\n"
        f"A password reset was requested for your account.\n\n"
        f"Click the link below to set a new password (valid for 1 hour):\n"
        f"{reset_url}\n\n"
        f"If you did not request this, ignore this email.\n\n"
        f"Plateau State University\nSmart Attendance System\n"
    )
    send_mail(
        subject, body,
        django_settings.DEFAULT_FROM_EMAIL,
        [user.email],
        fail_silently=True,
    )


@login_required
@role_required('admin')
def download_student_template(request):
    """Serve a blank CSV template for bulk student upload."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="plasu_students_template.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'first_name', 'last_name', 'email', 'matric_number',
        'department_code', 'level', 'phone', 'courses'
    ])
    # Two example rows
    writer.writerow(['Chidi', 'Okeke', 'chidi.okeke@student.plasu.edu.ng',
                     'PLASU/2023/FNAS/001', 'CSC', '200', '08012345678', 'CSC201,CSC203'])
    writer.writerow(['Amaka', 'Nwosu', '',
                     'PLASU/2023/ARTS/002', 'ENG', '100', '', 'ENG101'])
    return response